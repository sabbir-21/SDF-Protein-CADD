from tkinter import Tk, Canvas, Entry, Button, PhotoImage, StringVar, Radiobutton, filedialog, messagebox
from tkinter import *
import os, sys, time, requests, json, csv
from webbrowser import open_new_tab
from openpyxl import load_workbook, Workbook
from bs4 import BeautifulSoup
from urllib.parse import unquote


try:
    import pyi_splash
    pyi_splash.update_text('UI Loaded ...')
    pyi_splash.close()
except:
    pass


if hasattr(sys, '_MEIPASS'):
    image_folder = sys._MEIPASS
else:
    image_folder = os.path.dirname(__file__)
'''
image_folder = os.getcwd()+r'\assets'
'''

try:
    
    #response = requests.get('https://raw.githubusercontent.com/sabbir-21/Licence-manager/main/licence.txt')
    response = requests.get('https://gitlab.com/sabbir299/cadd-licence/-/raw/main/licence.txt')
    license = json.loads(response.text)
    ACTIVATION_CODE = license[os.getenv('COMPUTERNAME')+'_'+os.getenv('Username')]
    vv = license['version']
    '''
    ACTIVATION_CODE = "SABBIR-AHMED-0000-0001"
    vv = "Version: 1.6"
    '''

except Exception as e:
    import datetime
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(f"error.log", "a") as file:
        file.write(timestamp+'\n'+str(e)+'\n'+50* '='+'\n')
    exit()

FOLDER = r'C:\Key'
os.makedirs(FOLDER, exist_ok=True)
ACTIVATION_FILE = FOLDER+ r"\activation.txt"

'''
if os.path.exists('list.xlsx'):
    workbook = load_workbook('list.xlsx')
'''

window = Tk()
window.geometry("860x520")
window.title('CADD Helper | by Sabbir Ahmed')
window.resizable(0, 0)
photo = PhotoImage(file=os.path.join(image_folder, 'folder.png'))
window.iconphoto(False, photo)
window.configure(bg = "#3A7FF6")

#canvas = Canvas(window,bg = "#3A7FF6",height = 520,width = 860,bd = 0,highlightthickness = 0,relief = "ridge")
#canvas.place(x = 0, y = 0)
#================================Strings=============================
app_title = "Welcome to CADD Helper App"
app_description = "CADD Helper can be used to\ndownload sdf files from imppat\nand swisssimilarity and others\nonly one click."
app_version = "Version: 1.7"
#====================================================================

def callback(urlpersonal):
   open_new_tab(urlpersonal)
def open_popup():
    messagebox.showinfo("Developer Info", "Sabbir Ahmed\nPharmacy 6th batch\nMawlana Bhashani Science and Technology University\n\nhttps://www.facebook.com/sabbir299")

def mainApp():
    ###===================   main App =====================
    canvas = Canvas(window,bg = "#3A7FF6",height = 520,width = 860,bd = 0,highlightthickness = 0,relief = "ridge")
    canvas.place(x = 0, y = 0)
    canvas.create_rectangle(431,0,860,520,fill="#FFFFFF",outline="")
    #canvas.create_rectangle(431,0,1460,520,fill="#FFFFFF",outline="")
    #Left Side
    canvas.create_text(40,40,anchor="nw",text=app_title,fill="#FCFCFC",font=("Roboto Bold", -24))
    #canvas.create_text(140,100,anchor="nw",text="Select Option",fill="#FCFCFC",font=("Roboto Bold", -18))
    selected_option = StringVar(value="imppat")
    ver = canvas.create_text(40,380,anchor="nw",text=app_version,fill="#FCFCFC",font=("Actor Regular", -14))
    if vv==app_version:
        ver_check = canvas.create_text(140,380,anchor="nw",text="Latest version",fill="#FCFCFC",font=("Actor Regular", -14))
    else:
        ver_check = canvas.create_text(140,380,anchor="nw",text="Update available",fill="#FCFCFC",font=("Actor Regular", -14))
        def blink_update():
            canvas.itemconfigure(ver_check, state=HIDDEN if canvas.itemcget(ver_check, "state") == NORMAL else NORMAL)
            canvas.after(800, blink_update)
        blink_update()
    canvas.tag_bind(ver_check, "<Button-1>", lambda e: callback("https://github.com/sabbir-21/SDF-Protein-CADD"))
    #canvas.create_text(165,380,anchor="nw",text=VERSION,fill="#FCFCFC",font=("Actor Regular", 14 * -1))
    credit = canvas.create_text(40,407,anchor="nw",text="Developed by: SABBIR AHMED",fill="#FCFCFC",font=("Actor Regular", -20))
    canvas.tag_bind(credit, "<Button-1>", lambda e: open_popup())
    #canvas.tag_bind(credit, "<Enter>", lambda e: open_popup())
    #middle mouse
    canvas.tag_bind(credit, "<Button>", lambda e: open_popup())

    link = canvas.create_text(70,446,anchor="nw",text="https://facebook.com/sabbir299",fill="#FFFFFF",font=("Adamina Regular", -15))
    canvas.tag_bind(link, "<Button-1>", lambda e: callback("https://www.facebook.com/sabbir299"))
    #Right Side
    button_1 = Button(text='Select Excel sheet',command=lambda: choseExcel(), fg='white', bg='#3A7FF6', bd=0)
    button_1.place(x=565,y=133,width=146,height=36)

    csv_button = Button(text='CSV',command=lambda: chosecsv(), fg='white', bg='#3A7FF6', bd=0)
    csv_button.place(x=745,y=133,width=36,height=36)
    csv_button.place_forget()
    #csv_button.place(x=725,y=133,width=36,height=36)
    #button_1.config(state='disabled', bg="#FFFFFF", text="")

    entryone = canvas.create_text(492,190,anchor="nw",text="IMPPAT Plant URL",fill="#000000",font=("Actor Regular", -13))
    #under 2 line should be fixed later
    #canvas.main_image_1 = PhotoImage(file=os.path.join(image_folder, 'entry_1.png'))
    #canvas.create_image(618,232,image=canvas.main_image_1)
    #
    entry_1 = Entry(bd=0,bg="#F1F1F1",fg="#393b4f",highlightthickness=0)
    entry_1.place(x=491,y=209,width=324,height=44)

    #active_sheet = workbook.active
    #entry_1.insert(0, active_sheet)
    entry_1.delete(0, 'end')
    entry_1.insert(0, 'https://cb.imsc.res.in/imppat/phytochemical/')

    entrytwo = canvas.create_text(492,266,anchor="nw",text="Column Name containg ID/CID",fill="#000000",font=("Actor Regular", -13))
    #entrytwo_image = canvas.create_image(618,309,image=canvas.main_image_1)
    entry_2 = Entry(bd=0,bg="#F1F1F1",fg="#393b4f",highlightthickness=0)
    entry_2.place(x=491,y=286,width=294,height=44)
    entry_2.place_forget()
    #canvas.itemconfigure(entryone, text="IMPPAT Plant URL")
    canvas.itemconfigure(entrytwo, text="")
    #entry_2.insert(0, 'C')

    showText2 = canvas.create_text(492,356,anchor="nw",text="",fill="#000000",font=("Actor Regular", -13))
    showText = canvas.create_text(492,376,anchor="nw",text="",fill="#000000",font=("Actor Regular", -13))
    button_2 = Button(text='Download SDF',command=lambda: dl_impaat(), fg='white', bg='#3A7FF6', bd=0)
    button_2.place(x=565,y=412,width=146,height=36)

    def chosecsv():
        try:
            fileread=filedialog.askopenfilename(filetypes =[('CSV Files', '*.csv')])
            drugbank_ids = []
            with open(fileread, 'r') as csvfile:
                csvreader = csv.DictReader(csvfile, delimiter=';')
                for row in csvreader:
                    drugbank_ids.append(row['DrugBank ID'])

            # Create a new Excel workbook
            workbook = Workbook()
            worksheet = workbook.active

            # Write data to the Excel worksheet
            for row_index, drugbank_id in enumerate(drugbank_ids, start=1):
                worksheet.cell(row=row_index, column=1, value=drugbank_id)

            # Save the workbook to an Excel file
            workbook.save("csvToExcel.xlsx")
            canvas.itemconfigure(showText, text='Converted to csvToExcel.xlsx')
            canvas.update()
        except FileNotFoundError:
            pass
    #option if statement

    heading = canvas.create_text(530+10,40,anchor="nw",text='IMPPAT Medicinal',fill="#000000",font=("Roboto Bold", -24))
    heading2 = canvas.create_text(491+10,80,anchor="nw",text='Downloads SDF files from IMPPAT ID.\n[IMPPAT ID] to [SDF]',fill="#000000",font=("Roboto", -14))
    def update_selection():
        option = selected_option.get()
        if option == "swiss":
            canvas.itemconfigure(heading, text="Swiss Similarity")
            canvas.itemconfigure(heading2, text="Downloads SDF files from Drugbank ID.\n[Drugbank ID] to [SDF]")
            button_2 = Button(text='Download SDF',command=lambda: dl_swiss(), fg='white', bg='#3A7FF6', bd=0)
            button_2.place(x=545,y=412,width=146,height=36)
            button_1.place(x=545,y=133,width=146,height=36)
            csv_button.place(x=705,y=133,width=36,height=36)
            canvas.itemconfigure(entryone, text="Sheet Name")
            canvas.itemconfigure(entrytwo, text="Column Name containg ID/CID")
            entry_1.place(x=491,y=209,width=294,height=44)
            entry_2.place(x=491,y=286,width=294,height=44)
            entry_1.delete(0, 'end')
            entry_1.insert(0, 'ScreeningResults')
            entry_2.config(state='normal')
            entry_2.delete(0, 'end')
            entry_2.insert(0, 'A')
        elif option == "imppat":
            canvas.itemconfigure(heading, text='IMPPAT Medicinal')
            canvas.itemconfigure(heading2, text='Downloads SDF files from IMPPAT ID.\n[IMPPAT ID] to [SDF]')
            button_2 = Button(text='Download SDF',command=lambda: dl_impaat(), fg='white', bg='#3A7FF6', bd=0)
            button_2.place(x=545,y=412,width=146,height=36)
            entry_1.place(x=491,y=209,width=324,height=44)
            entry_2.place_forget()
            button_1.place_forget()
            csv_button.place_forget()
            canvas.itemconfigure(entryone, text="IMPPAT Plant URL")
            canvas.itemconfigure(entrytwo, text="")
            entry_1.delete(0, 'end')
            entry_1.insert(0, 'https://cb.imsc.res.in/imppat/phytochemical/')
            entry_2.delete(0, 'end')
            entry_2.insert(0, '')
            entry_2.config(state='disabled')
            
        elif option == "pubchem":
            canvas.itemconfigure(heading, text="Pubchem Library")
            canvas.itemconfigure(heading2, text="Downloads SDF files from PubChem CID.\n[PubChem CID] to [SDF]")
            button_2 = Button(text='Download SDF',command=lambda: dl_pubchem(), fg='white', bg='#3A7FF6', bd=0)
            button_2.place(x=545,y=412,width=146,height=36)
            button_1.place(x=545,y=133,width=146,height=36)
            entry_1.place(x=491,y=209,width=294,height=44)
            entry_2.place(x=491,y=286,width=294,height=44)
            csv_button.place_forget()
            canvas.itemconfigure(entryone, text="Sheet Name")
            canvas.itemconfigure(entrytwo, text="Column Name containg ID/CID")
            entry_1.delete(0, 'end')
            entry_1.insert(0, 'Sheet1')
            entry_2.config(state='normal')
            entry_2.delete(0, 'end')
            entry_2.insert(0, 'A')
        elif option == "smiletosdf":
            canvas.itemconfigure(heading, text="Pubchem SMILE to SDF")
            canvas.itemconfigure(heading2, text="Downloads SDF files from PubChem SMILES.\n[PubChem SMILES] to [SDF]")
            button_2 = Button(text='Download SDF',command=lambda: dl_smiletosdf(), fg='white', bg='#3A7FF6', bd=0)
            button_2.place(x=545,y=412,width=146,height=36)
            button_1.place(x=545,y=133,width=146,height=36)
            entry_1.place(x=491,y=209,width=294,height=44)
            entry_2.place(x=491,y=286,width=294,height=44)
            csv_button.place_forget()
            canvas.itemconfigure(entryone, text="Sheet Name")
            canvas.itemconfigure(entrytwo, text="Column Name containg ID/CID")
            entry_1.delete(0, 'end')
            entry_1.insert(0, 'Sheet1')
            entry_2.config(state='normal')
            entry_2.delete(0, 'end')
            entry_2.insert(0, 'A')
        elif option == "smile":
            canvas.itemconfigure(heading, text="Pubchem SMILES")
            canvas.itemconfigure(heading2, text="Downloads SMILES as txt from PubChem CID.\n[PubChem CID] to [SMILES]")
            button_2 = Button(text='Write SMILES',command=lambda: dl_smiles(), fg='white', bg='#3A7FF6', bd=0)
            button_2.place(x=545,y=412,width=146,height=36)
            button_1.place(x=545,y=133,width=146,height=36)
            entry_1.place(x=491,y=209,width=294,height=44)
            entry_2.place(x=491,y=286,width=294,height=44)
            csv_button.place_forget()
            canvas.itemconfigure(entryone, text="Sheet Name")
            canvas.itemconfigure(entrytwo, text="Column Name containg ID/CID")
            entry_1.delete(0, 'end')
            entry_1.insert(0, 'Sheet1')
            entry_2.config(state='normal')
            entry_2.delete(0, 'end')
            entry_2.insert(0, 'A')
        elif option == "img":
            canvas.itemconfigure(heading, text="2D Structure Download")
            canvas.itemconfigure(heading2, text="Downloads 2D PNG images from PubChem CID/IMPPAT/\nDrugbank. [ID] to [PNG]")
            button_2 = Button(text='Download 2D',command=lambda: dl_img(), fg='white', bg='#3A7FF6', bd=0)
            button_2.place(x=545,y=412,width=146,height=36)
            button_1.place(x=545,y=133,width=146,height=36)
            entry_1.place(x=491,y=209,width=294,height=44)
            entry_2.place(x=491,y=286,width=294,height=44)
            csv_button.place_forget()
            canvas.itemconfigure(entryone, text="Sheet Name")
            canvas.itemconfigure(entrytwo, text="Column Name containg ID/CID")
            entry_1.delete(0, 'end')
            entry_1.insert(0, 'Sheet1')
            entry_2.config(state='normal')
            entry_2.delete(0, 'end')
            entry_2.insert(0, 'A')
        elif option == "chem":
            canvas.itemconfigure(heading, text="Retrive Chemical Names")
            canvas.itemconfigure(heading2, text="Downloads Chemical names from PubChem CID.\n[PubChem CID] to [TXT]")
            button_2 = Button(text='Retrive Names',command=lambda: dl_chem(), fg='white', bg='#3A7FF6', bd=0)
            button_2.place(x=545,y=412,width=146,height=36)
            button_1.place(x=545,y=133,width=146,height=36)
            entry_1.place(x=491,y=209,width=294,height=44)
            entry_2.place(x=491,y=286,width=294,height=44)
            csv_button.place_forget()
            canvas.itemconfigure(entryone, text="Sheet Name")
            canvas.itemconfigure(entrytwo, text="Column Name containg ID/CID")
            entry_1.delete(0, 'end')
            entry_1.insert(0, 'Sheet1')
            entry_2.config(state='normal')
            entry_2.delete(0, 'end')
            entry_2.insert(0, 'A')

    #option0 = Radiobutton(window, text="Select Any Option", value="blankoption", variable=selected_option, command=print(""), fg='#000', bg='white', bd=0)
    #option0.place(x=120,y=132)
    option2 = Radiobutton(window, text="IMPPAT Medicinal [SDF]", value="imppat", variable=selected_option, command=lambda: update_selection(), fg='#000', bg='white', bd=0)
    option2.place(x=120,y=133)
    option1 = Radiobutton(window, text="Swiss Similarity (Drugbank) [SDF]", value="swiss", variable=selected_option, command=lambda: update_selection(), fg='#000', bg='white', bd=0)
    option1.place(x=120,y=161)
    option3 = Radiobutton(window, text="Pubchem CID to SDF [SDF]", value="pubchem", variable=selected_option, command=lambda: update_selection(), fg='#000', bg='white', bd=0)
    option3.place(x=120,y=189)
    option3 = Radiobutton(window, text="Pubchem SMILE to SDF [SDF]", value="smiletosdf", variable=selected_option, command=lambda: update_selection(), fg='#000', bg='white', bd=0)
    option3.place(x=120,y=217)
    option4 = Radiobutton(window, text="ADMET Accelerator (SMILES) [TXT]", value="smile", variable=selected_option, command=lambda: update_selection(), fg='#000', bg='white', bd=0)
    option4.place(x=120,y=245)
    option5 = Radiobutton(window, text="2D Structure Download [PNG]", value="img", variable=selected_option, command=lambda: update_selection(), fg='#000', bg='white', bd=0)
    option5.place(x=120,y=273)
    option6 = Radiobutton(window, text="Retrive Chemical Names [TXT]", value="chem", variable=selected_option, command=lambda: update_selection(), fg='#000', bg='white', bd=0)
    option6.place(x=120,y=301)

    def choseExcel():
        try:
            fileread=filedialog.askopenfilename(filetypes =[('Excel Files', '*.xlsx;*.xls')])
            tf = open(f'{fileread}', 'rb')
            global workbook, active_sheet
            workbook = load_workbook(tf)
            active_sheet = workbook.active.title
            entry_1.delete(0, 'end')
            entry_1.insert(0, active_sheet)
        except FileNotFoundError:
            pass
        
    def dl_impaat():
        try:
            #bl_on()
            sheet = entry_1.get()
            r = requests.get(sheet)
            soup = BeautifulSoup(r.text, 'html.parser')
            table = soup.find('table', class_ ="phytochem table").find('tbody').find_all('tr')
            canvas.itemconfigure(showText2, text=f"Total files with duplicate: {len(table)}")
            canvas.update()
            canvas.itemconfigure(showText, text='Download Started...')
            canvas.update()
            folder_ext = unquote(sheet.split("/")[-1])
            os.makedirs('SDF_Imppat_'+folder_ext, exist_ok=True)
            i = 0
            j = 0
            for row in table:
                val = row.find_all('td')[2].text
                URL = f'https://cb.imsc.res.in/imppat/images/3D/SDF/{val}_3D.sdf'
                name = URL.split('/')[-1]
                r = requests.get(URL)
                r_dl = r.content
                if r.status_code == 200:
                    open(f"SDF_Imppat_{folder_ext}/{name}", "wb+").write(r_dl)
                else:
                    j += 1
                    canvas.itemconfigure(showText, text=f'Failed {j}')
                    canvas.update()
                    with open(f"failed_imppat_{folder_ext}.txt", "a") as file:
                        file.write(str(val)+'\n')
                    continue
                i+=1
                canvas.itemconfigure(showText, text=f'{i} sdf file Downloaded')
                canvas.update()
               
            canvas.itemconfigure(showText, text=f'Download Completed {i} sdf Files & failed {j}')
        except Exception as e:
            canvas.itemconfigure(showText, text=e)
            canvas.update()
        '''
        except ConnectionError:
            canvas.itemconfigure(showText, text='No Internet Connection')
            canvas.update()
        '''
        #canvas.update()
    def dl_swiss():
        try:
            sheet = entry_1.get()
            column = entry_2.get()
            worksheet = workbook[sheet]
            num_rows = worksheet.max_row
            #worksheet = workbook.active
            canvas.itemconfigure(showText2, text=f'Expected Download: {num_rows}')
            canvas.update()
            canvas.itemconfigure(showText, text='Download Started...')
            canvas.update()
            folder_ext = str(round(time.time()))
            os.makedirs('SDF_Swiss_'+folder_ext, exist_ok=True)
            i = 0
            j = 0
            for cell in worksheet[column]:
                val = str(cell.value)
                URL = f'https://go.drugbank.com/structures/small_molecule_drugs/{val}.sdf'
                name = URL.split('/')[-1]
                r = requests.get(URL)
                r_dl = r.content
                if r.status_code == 200:
                    open(f"SDF_Swiss_{folder_ext}/{name}", "wb+").write(r_dl)
                else:
                    j += 1
                    canvas.itemconfigure(showText, text=f'Failed {j}')
                    canvas.update()
                    with open(f"failed_swiss.txt", "a") as file:
                        file.write(str(val)+'\n')
                    continue
                i+=1
                canvas.itemconfigure(showText, text=f'{i} sdf file Downloaded')
                canvas.update()
                #print(f'{i} sdf file Downloaded')
            canvas.itemconfigure(showText, text=f'Download Completed {i} sdf Files & failed {j}')
        except NameError:
            canvas.itemconfigure(showText, text='Excel File not selected')
            canvas.update()
    def dl_smiles():
        try:
            sheet = entry_1.get()
            column = entry_2.get()
            worksheet = workbook[sheet]
            num_rows = worksheet.max_row
            canvas.itemconfigure(showText2, text=f'Expected Download: {num_rows}')
            canvas.itemconfigure(showText, text='Process Started...')
            canvas.update()
            base_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"
            folder_ext = str(round(time.time()))
            i = 0
            j = 0
            with open(f"all_smiles_{folder_ext}.txt", "a") as file:
                file.write("\n")
            for cell in worksheet[column]:
                val = str(cell.value)
                smiles_endpoint = f"{base_url}/cid/{val}/property/CanonicalSMILES/json"
                response = requests.get(smiles_endpoint)
                if response.status_code == 200:
                    response_list = response.json()
                    canonical_smiles = response_list["PropertyTable"]["Properties"][0]["CanonicalSMILES"]
                    with open(f"all_smiles_{folder_ext}.txt", "a") as file:
                        file.write(canonical_smiles + "\n")
                else:
                    j += 1
                    canvas.itemconfigure(showText, text=f'Failed {j}')
                    canvas.update()
                    with open(f"failed_smiles.txt", "a") as file:
                        file.write(str(val)+'\n')
                    continue
                i +=1
                canvas.itemconfigure(showText, text=f'{i} SMILES written')
                canvas.update()
            canvas.itemconfigure(showText, text=f'Write Completed {i} Canonical SMILES & failed {j}')
        except NameError:
            canvas.itemconfigure(showText, text='Excel File not selected')
            canvas.update()
    def dl_pubchem():
        try:
            sheet = entry_1.get()
            column = entry_2.get()
            worksheet = workbook[sheet]
            num_rows = worksheet.max_row
            canvas.itemconfigure(showText2, text=f'Expected Download: {num_rows}')
            canvas.itemconfigure(showText, text='Download Started...')
            canvas.update()
            folder_ext = str(round(time.time()))
            base_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"
            os.makedirs('SDF_Pubchem_'+folder_ext, exist_ok=True)
            i = 0
            j = 0
            for cell in worksheet[column]:
                val = str(cell.value)
                pub_endpoint = f"{base_url}/cid/{val}/record/SDF/?record_type=3d&response_type=save&response_basename=Conformer3D_CID_{val}"
                response = requests.get(pub_endpoint)
                response_dl = response.content
                
                if response.status_code == 200:
                    with open(f"SDF_Pubchem_{folder_ext}/Conformer3D_CID_{val}.sdf", "wb") as f:
                        f.write(response_dl)
                else:
                    twod = f"{base_url}/cid/{val}/record/SDF/?record_type=2d&response_type=save&response_basename=Conformer2D_CID_{val}"
                    twod_response = requests.get(twod)
                    twod_dl = twod_response.content
                    if twod_response.status_code == 200:
                        with open(f"SDF_Pubchem_{folder_ext}/Conformer2D_CID_{val}.sdf", "wb") as f:
                            f.write(twod_dl)
                    else:
                        j += 1
                        canvas.itemconfigure(showText, text=f'Failed {j}')
                        canvas.update()
                        with open(f"failed_pubchem.txt", "a") as file:
                            file.write(str(val)+'\n')
                        continue
                i +=1
                canvas.itemconfigure(showText, text=f'{i} sdf file Downloaded')
                canvas.update()
            canvas.itemconfigure(showText, text=f'Download Completed {i} sdf Files & failed {j}')
        except NameError:
            canvas.itemconfigure(showText, text='Excel File not selected')
            canvas.update()
    #start
    def dl_smiletosdf():
        try:
            sheet = entry_1.get()
            column = entry_2.get()
            worksheet = workbook[sheet]
            num_rows = worksheet.max_row
            canvas.itemconfigure(showText2, text=f'Expected Download: {num_rows}')
            canvas.itemconfigure(showText, text='Download Started...')
            canvas.update()
            folder_ext = str(round(time.time()))
            base_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"
            os.makedirs('Smile_to_SDF_'+folder_ext, exist_ok=True)
            i = 0
            j = 0
            for cell in worksheet[column]:
                val = str(cell.value)
                cid_url = f"{base_url}/smiles/{val}/cids/TXT"
                cid_response = requests.get(cid_url)
                cid = cid_response.text.strip()
                if cid_response.status_code != 200:
                    j += 1
                    canvas.itemconfigure(showText, text=f'Failed {j}')
                    canvas.update()
                    with open(f"failed_smileToSdf_{folder_ext}.txt", "a") as file:
                        file.write(str(val)+'\n\n')
                    continue
                pub_endpoint = f"{base_url}/cid/{cid}/record/SDF/?record_type=3d&response_type=save&response_basename=Conformer3D_CID_{cid}"
                response = requests.get(pub_endpoint)
                response_dl = response.content
                
                if response.status_code == 200:
                    with open(f"Smile_to_SDF_{folder_ext}/Conformer3D_CID_{cid}.sdf", "wb") as f:
                        f.write(response_dl)
                else:
                    twod = f"{base_url}/cid/{cid}/record/SDF/?record_type=2d&response_type=save&response_basename=Conformer2D_CID_{cid}"
                    twod_response = requests.get(twod)
                    twod_dl = twod_response.content
                    if twod_response.status_code == 200:
                        with open(f"Smile_to_SDF_{folder_ext}/Conformer2D_CID_{cid}.sdf", "wb") as f:
                            f.write(twod_dl)
                    else:
                        continue
                i +=1
                canvas.itemconfigure(showText, text=f'{i} sdf file Downloaded')
                canvas.update()
            canvas.itemconfigure(showText, text=f'Download Completed {i} sdf Files & failed {j}')
        except NameError:
            canvas.itemconfigure(showText, text='Excel File not selected')
            canvas.update()

    def dl_img():
        try:
            sheet = entry_1.get()
            column = entry_2.get()
            worksheet = workbook[sheet]
            num_rows = worksheet.max_row
            canvas.itemconfigure(showText2, text=f'Expected Download: {num_rows}')
            canvas.itemconfigure(showText, text='Download Started...')
            canvas.update()
            folder_ext = str(round(time.time()))
            os.makedirs('2D-Structure_'+folder_ext, exist_ok=True)
            i = 0
            j = 0
            for cell in worksheet[column]:
                val = str(cell.value)
                if "IMPHY" in val:
                    img_url_api = f"https://cb.imsc.res.in/imppat/images/2D_IMAGE/PNG/{val}.png"
                elif "DB" in val:
                    img_url_api = f"https://go.drugbank.com/structures/{val}/image.png"
                elif "CHEMBL" in val:
                    img_url_api = f"https://pubchem.ncbi.nlm.nih.gov/image/imagefly.cgi?cid={val}&width=500&height=500"
                else:
                    img_url_api = f"https://pubchem.ncbi.nlm.nih.gov/image/imagefly.cgi?cid={val}&width=500&height=500"
                response = requests.get(img_url_api)
                response_dl = response.content
                
                if response.status_code == 200:
                    with open(f"2D-Structure_{folder_ext}/{val}.png", "wb") as f:
                        f.write(response_dl)
                else:
                    j += 1
                    canvas.itemconfigure(showText, text=f'Failed {j}')
                    canvas.update()
                    with open(f"failed_2D-Structure{folder_ext}.txt", "a") as file:
                        file.write(str(val)+'\n')
                    continue
                i +=1
                canvas.itemconfigure(showText, text=f'{i} png file Downloaded')
                canvas.update()
            canvas.itemconfigure(showText, text=f'Download Completed {i} png images & failed {j}')
        except NameError:
            canvas.itemconfigure(showText, text='Excel File not selected')
            canvas.update()
    def dl_chem():
        try:
            sheet = entry_1.get()
            column = entry_2.get()
            worksheet = workbook[sheet]
            num_rows = worksheet.max_row
            canvas.itemconfigure(showText2, text=f'Expected Download: {num_rows}')
            canvas.itemconfigure(showText, text='Process Started...')
            canvas.update()
            base_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"
            folder_ext = str(round(time.time()))
            i = 0
            j = 0
            #with open(f"chemical_names_{folder_ext}.txt", "w") as file:
            for cell in worksheet[column]:
                val = str(cell.value)
                chem_endpoint = f"{base_url}/cid/{val}/property/IUPACName/json"
                response = requests.get(chem_endpoint)
                if response.status_code == 200:
                    response_list = response.json()
                    chemical_name = response_list["PropertyTable"]["Properties"][0]["IUPACName"]
                    with open(f"chemical_names_{folder_ext}.txt", "a") as file:
                        file.write(chemical_name + "\n\n")
                else:
                    j += 1
                    canvas.itemconfigure(showText, text=f'Failed {j}')
                    canvas.update()
                    with open(f"failed_chemical_names_{folder_ext}.txt", "a") as file:
                        file.write(str(val)+'\n')
                    continue
                i +=1
                canvas.itemconfigure(showText, text=f'{i} Chemical Names written')
                canvas.update()
            canvas.itemconfigure(showText, text=f'Write Completed {i} Chemical Names & failed {j}')
        except NameError:
            canvas.itemconfigure(showText, text='Excel File not selected')
            canvas.update()
    #####=====================end main app

###===================  Start Check Activation   ================================================================================

def check_activation():
    entered_code = entry_2.get().strip()
    if entered_code == ACTIVATION_CODE:
        with open(ACTIVATION_FILE, "w") as f:
            f.write(ACTIVATION_CODE)
        canvas_s.destroy()
        entry_1.destroy()
        entry_2.destroy()
        button_1.destroy()
        mainApp()
    else:
        entry_2.delete(0, 'end')
        entry_2.insert(0, 'Invalid Licence. Please try again.')

if os.path.exists(ACTIVATION_FILE):
    with open(ACTIVATION_FILE, "r") as f:
        saved_code = f.read().strip()
    if saved_code == ACTIVATION_CODE:
        mainApp()
    else:
        #start
        canvas_s = Canvas(window,bg = "#3A7FF6",height = 520,width = 860,bd = 0,highlightthickness = 0,relief = "ridge")
        canvas_s.place(x = 0, y = 0)
        canvas_s.create_rectangle(431,0,860,520,fill="#FFFFFF",outline="")
        #Left Side
        canvas_s.create_text(40,68,anchor="nw",text=app_title,fill="#FCFCFC",font=("Roboto Bold", -24))
        #110-115 height, 40 start X, 120 length Y
        canvas_s.create_rectangle(40,110,120,115,fill="#FCFCFC",outline="")
        canvas_s.create_text(40,132,anchor="nw",text=app_description,fill="#FCFCFC",font=("Actor Regular", -24))

        ver = canvas_s.create_text(40,380,anchor="nw",text=app_version,fill="#FCFCFC",font=("Actor Regular", -14))
        canvas_s.tag_bind(ver, "<Button-1>", lambda e: callback("https://github.com/sabbir-21/SDF-Protein-CADD"))
        credit = canvas_s.create_text(40,407,anchor="nw",text="Developed by: SABBIR AHMED",fill="#FCFCFC",font=("Actor Regular", -20))
        canvas_s.tag_bind(credit, "<Button-1>", lambda e: open_popup())
        canvas_s.tag_bind(credit, "<Button>", lambda e: open_popup())
        link = canvas_s.create_text(70,446,anchor="nw",text="https://facebook.com/sabbir299",fill="#FFFFFF",font=("Adamina Regular", -15))
        canvas_s.tag_bind(link, "<Button-1>", lambda e: callback("https://www.facebook.com/sabbir299"))
        #Right Side
        canvas_s.create_text(482,68.0,anchor="nw",text="Enter Your Licence Key",fill="#505485",font=("Roboto Bold", -24))
        canvas_s.create_text(482,120,anchor="nw",text="Send the Device ID to developer to get licence Key",fill="#505485",font=("Actor Regular", 13 * -1))
        canvas_s.create_text(492,160,anchor="nw",text="Your Device ID",fill="#505485",font=("Actor Regular", -13))

        entry_image_1 = PhotoImage(file=os.path.join(image_folder, 'entry_1.png'))
        entry_bg_1 = canvas_s.create_image(618,202,image=entry_image_1)
        entry_1 = Entry(bd=0,bg="#F1F1F1",fg="#000716",highlightthickness=0)
        entry_1.place(x=491,y=179,width=254,height=44)
        ID = os.getenv('COMPUTERNAME')+'_'+os.getenv('UserName')
        entry_1.insert(0, ID)

        canvas_s.create_text(492,266,anchor="nw",text="Licence Key",fill="#505485",font=("Actor Regular", -13))
        entry_bg_2 = canvas_s.create_image(618,309,image=entry_image_1)
        entry_2 = Entry(bd=0,bg="#F1F1F1",fg="#000716",highlightthickness=0)
        entry_2.place(x=491,y=286,width=254,height=44.0)

        button_image_1 = PhotoImage(file=os.path.join(image_folder, 'button_1.png'))
        button_1 = Button(image=button_image_1,borderwidth=0,highlightthickness=0,command=lambda: check_activation(),relief="flat")
        button_1.place(x=545,y=392,width=146.0,height=36.0)

        #copy option of device id
        def copy_text():
            selected_text = entry_1.get()
            canvas_s.clipboard_clear()
            canvas_s.clipboard_append(selected_text)
        context_menu = Menu(canvas_s, tearoff=0)
        context_menu.add_command(label="Copy Device ID", command=copy_text)
        def show_context_menu(event):
            context_menu.post(event.x_root, event.y_root)
        entry_1.bind("<Button-3>", show_context_menu)
        #end copy
else:
    #start
    canvas_s = Canvas(window,bg = "#3A7FF6",height = 520,width = 860,bd = 0,highlightthickness = 0,relief = "ridge")
    canvas_s.place(x = 0, y = 0)
    canvas_s.create_rectangle(431,0,860,520,fill="#FFFFFF",outline="")
    #Left Side
    canvas_s.create_text(40,68,anchor="nw",text=app_title,fill="#FCFCFC",font=("Roboto Bold", -24))
    #110-115 height, 40 start X, 120 length Y
    canvas_s.create_rectangle(40,110,120,115,fill="#FCFCFC",outline="")
    canvas_s.create_text(40,132,anchor="nw",text=app_description,fill="#FCFCFC",font=("Actor Regular", -24))

    ver = canvas_s.create_text(40,380,anchor="nw",text=app_version,fill="#FCFCFC",font=("Actor Regular", -14))
    canvas_s.tag_bind(ver, "<Button-1>", lambda e: callback("https://github.com/sabbir-21/SDF-Protein-CADD"))
    credit = canvas_s.create_text(40,407,anchor="nw",text="Developed by: SABBIR AHMED",fill="#FCFCFC",font=("Actor Regular", -20))
    canvas_s.tag_bind(credit, "<Button-1>", lambda e: open_popup())
    canvas_s.tag_bind(credit, "<Button>", lambda e: open_popup())
    link = canvas_s.create_text(70,446,anchor="nw",text="https://facebook.com/sabbir299",fill="#FFFFFF",font=("Adamina Regular", -15))
    canvas_s.tag_bind(link, "<Button-1>", lambda e: callback("https://www.facebook.com/sabbir299"))
    #Right Side
    canvas_s.create_text(482,68,anchor="nw",text="Enter Your Licence Key",fill="#505485",font=("Roboto Bold", -24))
    canvas_s.create_text(482,120,anchor="nw",text="Send the Device ID to developer to get licence Key",fill="#505485",font=("Actor Regular", 13 * -1))
    canvas_s.create_text(492,160,anchor="nw",text="Your Device ID",fill="#505485",font=("Actor Regular", -13))

    entry_image_1 = PhotoImage(file=os.path.join(image_folder, 'entry_1.png'))
    entry_bg_1 = canvas_s.create_image(618,202,image=entry_image_1)
    entry_1 = Entry(bd=0,bg="#F1F1F1",fg="#000716",highlightthickness=0)
    entry_1.place(x=491,y=179,width=254,height=44)
    ID = os.getenv('COMPUTERNAME')+'_'+os.getenv('UserName')
    entry_1.insert(0, ID)

    canvas_s.create_text(492,266,anchor="nw",text="Licence Key",fill="#505485",font=("Actor Regular", -13))
    entry_bg_2 = canvas_s.create_image(618,309,image=entry_image_1)
    entry_2 = Entry(bd=0,bg="#F1F1F1",fg="#000716",highlightthickness=0)
    entry_2.place(x=491,y=286,width=254,height=44)

    button_image_1 = PhotoImage(file=os.path.join(image_folder, 'button_1.png'))
    button_1 = Button(image=button_image_1,borderwidth=0,highlightthickness=0,command=lambda: check_activation(),relief="flat")
    button_1.place(x=545,y=392,width=146,height=36)
    #copy option of device id
    def copy_text():
        selected_text = entry_1.get()
        canvas_s.clipboard_clear()
        canvas_s.clipboard_append(selected_text)
    context_menu = Menu(canvas_s, tearoff=0)
    context_menu.add_command(label="Copy Device ID", command=copy_text)
    def show_context_menu(event):
        context_menu.post(event.x_root, event.y_root)
    entry_1.bind("<Button-3>", show_context_menu)
    #end copy
    #end
###===================== End Check Activvation   ===========================

window.mainloop()
